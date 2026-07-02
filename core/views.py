from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
import qrcode
import json
import base64
import io
from datetime import timedelta

from .models import (
    Automobile, Region, Utilisateur, RoleUtilisateur,
    StatutVignette, StatutVignetteChoix, StatutApprobationVehicule, StatutPhysiqueVignette,
    TypeModification, CodeSecurite, HistoriqueConsultation,
    Paiement, StatutPaiement, OperateurPaiement, OTPVerification,
    ParametrePlateforme, DemandeTransfert, StatutTransfert,
    Plainte, StatutPlainte, NotificationLog, JournalAudit, CategorieAudit,
)
from .sms_service import sms_confirmation_paiement, sms_otp, sms_plainte_recue, sms_transfert_approuve
import random
from .serializers import (
    AutomobileReadSerializer, AutomobileWriteSerializer,
    RegionSerializer, UtilisateurSerializer, UtilisateurCreateSerializer,
    UtilisateurUpdateSerializer, ParametrePlateformeSerializer,
    StatutVignetteSerializer, CodeSecuriteSerializer, HistoriqueConsultationSerializer,
    PaiementSerializer, DemandeTransfertSerializer, JournalAuditSerializer,
)
from .permissions import AdminOnlyPermission, RoleBasedPermission


def log_audit(request, categorie, action, detail='', objet_type='', objet_id='', objet_label=''):
    u = request.user if request.user and request.user.is_authenticated else None
    JournalAudit.objects.create(
        utilisateur=u,
        utilisateur_email=u.email if u else '',
        utilisateur_role=u.role if u else '',
        categorie=categorie,
        action=action,
        detail=detail,
        objet_type=objet_type,
        objet_id=str(objet_id) if objet_id else '',
        objet_label=objet_label,
        ip_address=request.META.get('REMOTE_ADDR'),
    )


# =============== AUTH ===============

class MyCarTokenSerializer(TokenObtainPairSerializer):
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)
        token['role']       = user.role
        token['nom']        = user.nom
        token['prenom']     = user.prenom
        token['email']      = user.email
        token['region_id']  = user.region_id
        token['region_nom'] = user.region.nom_region if user.region else None
        return token

class MyCarTokenObtainPairView(TokenObtainPairView):
    serializer_class = MyCarTokenSerializer

    def post(self, request, *args, **kwargs):
        response = super().post(request, *args, **kwargs)
        if response.status_code == 200:
            try:
                from .models import Utilisateur
                u = Utilisateur.objects.get(email=request.data.get('email', ''))
                JournalAudit.objects.create(
                    utilisateur=u,
                    utilisateur_email=u.email,
                    utilisateur_role=u.role,
                    categorie=CategorieAudit.CONNEXION,
                    action='Connexion au système',
                    ip_address=request.META.get('REMOTE_ADDR'),
                )
            except Exception:
                pass
        return response

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def me(request):
    serializer = UtilisateurSerializer(request.user)
    return Response(serializer.data)


class UtilisateurViewSet(viewsets.ModelViewSet):
    queryset = Utilisateur.objects.all()
    permission_classes = [AdminOnlyPermission]

    def get_serializer_class(self):
        if self.action == 'create':
            return UtilisateurCreateSerializer
        return UtilisateurSerializer


class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer
    permission_classes = [IsAuthenticated]


class AutomobileViewSet(viewsets.ModelViewSet):
    queryset = Automobile.objects.select_related('region', 'statut_actuel').all()
    permission_classes = [RoleBasedPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['immatriculation', 'nom', 'prenom', 'numero_chassis', 'marque']
    ordering_fields = ['date_creation', 'immatriculation']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return AutomobileWriteSerializer
        return AutomobileReadSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user

        # Agent DGI : restreint à sa région
        if user.role == RoleUtilisateur.AGENT_DGI and user.region_id:
            qs = qs.filter(region_id=user.region_id)

        region = self.request.query_params.get('region')
        statut = self.request.query_params.get('statut')
        type_vehicule = self.request.query_params.get('type_vehicule')
        if region:
            qs = qs.filter(region__nom_region__iexact=region)
        if statut:
            qs = qs.filter(statut_actuel__statut__iexact=statut)
        if type_vehicule:
            qs = qs.filter(type_vehicule__iexact=type_vehicule)
        return qs

    @action(detail=True, methods=['get'])
    def generer_qr(self, request, pk=None):
        automobile = self.get_object()
        qr_data = automobile.generer_qr_data()
        qr = qrcode.QRCode(
            version=10,
            error_correction=qrcode.constants.ERROR_CORRECT_H,
            box_size=10,
            border=4,
        )
        qr.add_data(json.dumps(qr_data))
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white")
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        qr_base64 = base64.b64encode(buffer.getvalue()).decode()

        return Response({'qr_code_base64': qr_base64, 'qr_data': qr_data})

    @action(detail=True, methods=['get'])
    def statuts(self, request, pk=None):
        """Historique des statuts d'un véhicule"""
        automobile = self.get_object()
        statuts = automobile.statuts.order_by('-date_creation')
        serializer = StatutVignetteSerializer(statuts, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def paiements(self, request, pk=None):
        """Liste des paiements d'un véhicule"""
        automobile = self.get_object()
        paiements = Paiement.objects.filter(automobile=automobile).order_by('-date_initiation')
        serializer = PaiementSerializer(paiements, many=True)
        return Response(serializer.data)


class StatutVignetteViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = StatutVignette.objects.select_related('automobile', 'operateur').all()
    serializer_class = StatutVignetteSerializer
    permission_classes = [RoleBasedPermission]


class CodeSecuriteViewSet(viewsets.ModelViewSet):
    queryset = CodeSecurite.objects.select_related('automobile', 'generateur').all()
    serializer_class = CodeSecuriteSerializer
    permission_classes = [RoleBasedPermission]

    def perform_create(self, serializer):
        serializer.save(generateur=self.request.user)


class HistoriqueConsultationViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = HistoriqueConsultation.objects.select_related('utilisateur', 'automobile').all()
    serializer_class = HistoriqueConsultationSerializer
    permission_classes = [RoleBasedPermission]
    filter_backends = [filters.OrderingFilter]
    ordering = ['-date_consultation']


class JournalAuditViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = JournalAuditSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        roles_autorises = {RoleUtilisateur.ADMIN_SYS, RoleUtilisateur.SUP_DGI}
        if self.request.user.role not in roles_autorises:
            return JournalAudit.objects.none()
        qs = JournalAudit.objects.all()
        categorie = self.request.query_params.get('categorie')
        user_email = self.request.query_params.get('utilisateur_email')
        date_debut = self.request.query_params.get('date_debut')
        date_fin = self.request.query_params.get('date_fin')
        if categorie:
            qs = qs.filter(categorie=categorie)
        if user_email:
            qs = qs.filter(utilisateur_email__icontains=user_email)
        if date_debut:
            qs = qs.filter(date_action__date__gte=date_debut)
        if date_fin:
            qs = qs.filter(date_action__date__lte=date_fin)
        return qs


# =============== CODE SÉCURITÉ ===============

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generer_code_securite(request, automobile_id):
    """Génère un CodeSecurite usage unique pour un véhicule (Agent DGI / Superviseur / Admin)."""
    roles_autorises = {
        RoleUtilisateur.AGENT_DGI,
        RoleUtilisateur.SUP_DGI,
        RoleUtilisateur.ADMIN_SYS,
    }
    if request.user.role not in roles_autorises:
        return Response({'error': 'Accès non autorisé.'}, status=403)

    try:
        auto = Automobile.objects.get(id=automobile_id)
    except Automobile.DoesNotExist:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    # Invalider les codes actifs précédents pour ce véhicule
    CodeSecurite.objects.filter(automobile=auto, statut_usage='ACTIF').update(statut_usage='UTILISE')

    code = CodeSecurite.objects.create(automobile=auto, generateur=request.user)

    return Response({
        'code':        code.code,
        'automobile':  auto.immatriculation,
        'generateur':  f"{request.user.nom} {request.user.prenom}",
        'expire_apres': '30 minutes (usage unique)',
    }, status=201)


# =============== STATUTS ===============

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def modifier_statut(request):
    """
    Modification manuelle du statut (Admin système uniquement).
    Transitions autorisées : ROUGE → VERT ou ROUGE → ORANGE.
    Champs obligatoires : code_securite, numero_recu, notes_admin.
    """
    if request.user.role != RoleUtilisateur.ADMIN_SYS:
        return Response({'error': 'Seul l\'administrateur système peut modifier manuellement le statut.'}, status=403)

    immatriculation = request.data.get('immatriculation')
    region_nom      = request.data.get('region')
    code            = request.data.get('code_securite')
    nouveau_statut  = request.data.get('nouveau_statut')
    numero_recu     = request.data.get('numero_recu', '').strip()
    notes_admin     = request.data.get('notes_admin', '').strip()

    if not all([immatriculation, region_nom, code, nouveau_statut, numero_recu, notes_admin]):
        return Response({
            'error': 'Champs obligatoires : immatriculation, region, code_securite, '
                     'nouveau_statut, numero_recu, notes_admin'
        }, status=400)

    # Seules transitions autorisées : ROUGE → VERT ou ROUGE → ORANGE
    transitions_autorisees = {
        StatutVignetteChoix.ROUGE: [StatutVignetteChoix.VERT, StatutVignetteChoix.ORANGE],
    }

    try:
        auto = Automobile.objects.get(
            immatriculation=immatriculation,
            region__nom_region__iexact=region_nom,
        )
    except Automobile.DoesNotExist:
        return Response({'error': 'Véhicule non trouvé'}, status=404)

    statut_courant = auto.statut_actuel.statut if auto.statut_actuel else StatutVignetteChoix.ROUGE

    if statut_courant not in transitions_autorisees:
        return Response({
            'error': f'Transition non autorisée. Le véhicule est actuellement {statut_courant}. '
                     f'Seul un véhicule ROUGE peut être modifié manuellement.'
        }, status=400)

    if nouveau_statut not in transitions_autorisees[statut_courant]:
        return Response({
            'error': f'Transition {statut_courant} → {nouveau_statut} non autorisée. '
                     f'Transitions valides depuis ROUGE : VERT ou ORANGE.'
        }, status=400)

    try:
        code_obj = CodeSecurite.objects.get(automobile=auto, code=code, statut_usage='ACTIF')
    except CodeSecurite.DoesNotExist:
        return Response({'error': 'Code sécurité invalide ou déjà utilisé'}, status=403)

    statut = StatutVignette.objects.create(
        automobile=auto,
        statut=nouveau_statut,
        date_debut_validite=timezone.now().date(),
        date_fin_validite=timezone.now().date().replace(month=12, day=31),
        type_modification=TypeModification.MANUELLE,
        operateur=request.user,
        numero_recu=numero_recu,
        notes_admin=notes_admin,
    )
    auto.statut_actuel = statut
    auto.save(update_fields=['statut_actuel'])

    code_obj.statut_usage = 'UTILISE'
    code_obj.date_utilisation = timezone.now()
    code_obj.save(update_fields=['statut_usage', 'date_utilisation'])

    # Traçabilité audit
    HistoriqueConsultation.objects.create(
        utilisateur=request.user,
        automobile=auto,
        action_performee=f'Transition manuelle {statut_courant} → {nouveau_statut} | Reçu: {numero_recu}',
        ip_address=request.META.get('REMOTE_ADDR'),
    )

    return Response({
        'success': True,
        'ancien_statut': statut_courant,
        'nouveau_statut': statut.statut,
        'numero_recu': numero_recu,
    })


# =============== APPROBATION VÉHICULE ===============

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def approuver_vehicule(request, automobile_id):
    """
    Approuver ou rejeter un véhicule (Admin système uniquement).
    Body: { "decision": "APPROUVE"|"REJETE"|"SUSPENDU", "notes": "..." }
    Lors de l'approbation, le statut vignette initial est automatiquement ROUGE.
    """
    if request.user.role != RoleUtilisateur.ADMIN_SYS:
        return Response({'error': 'Seul l\'administrateur système peut approuver les véhicules.'}, status=403)

    decision = request.data.get('decision', '').strip().upper()
    notes    = request.data.get('notes', '').strip()

    if decision not in StatutApprobationVehicule.values:
        return Response({'error': f'Décision invalide. Valeurs : {StatutApprobationVehicule.values}'}, status=400)

    try:
        auto = Automobile.objects.get(id=automobile_id)
    except Automobile.DoesNotExist:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    auto.statut_approbation = decision
    auto.notes_approbation  = notes
    auto.save(update_fields=['statut_approbation', 'notes_approbation'])

    # À l'approbation : créer le statut ROUGE initial si aucun statut n'existe
    if decision == StatutApprobationVehicule.APPROUVE and not auto.statut_actuel:
        sv = StatutVignette.objects.create(
            automobile=auto,
            statut=StatutVignetteChoix.ROUGE,
            date_debut_validite=timezone.now().date(),
            date_fin_validite=timezone.now().date().replace(month=12, day=31),
            type_modification=TypeModification.MANUELLE,
            operateur=request.user,
            notes_admin='Statut initial ROUGE à l\'approbation du véhicule.',
        )
        auto.statut_actuel = sv
        auto.save(update_fields=['statut_actuel'])

    HistoriqueConsultation.objects.create(
        utilisateur=request.user,
        automobile=auto,
        action_performee=f'Décision approbation : {decision} | Notes : {notes}',
        ip_address=request.META.get('REMOTE_ADDR'),
    )
    log_audit(request, CategorieAudit.VEHICULE,
              f'Approbation véhicule : {decision}',
              detail=notes,
              objet_type='Automobile', objet_id=auto.id, objet_label=auto.immatriculation)

    return Response({
        'success': True,
        'immatriculation': auto.immatriculation,
        'statut_approbation': auto.statut_approbation,
        'statut_vignette': auto.statut_actuel.statut if auto.statut_actuel else None,
    })


# =============== DASHBOARD ===============

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_superviseur(request):
    qs = Automobile.objects.all()
    if request.user.role == RoleUtilisateur.AGENT_DGI and request.user.region_id:
        qs = qs.filter(region_id=request.user.region_id)

    stats = qs.aggregate(
        total_verts=Count('id', filter=Q(statut_actuel__statut='VERT')),
        total_oranges=Count('id', filter=Q(statut_actuel__statut='ORANGE')),
        total_rouges=Count('id', filter=Q(statut_actuel__statut='ROUGE')),
        total=Count('id'),
        # Approbation
        en_attente_approbation=Count('id', filter=Q(statut_approbation='EN_ATTENTE')),
        approuves=Count('id', filter=Q(statut_approbation='APPROUVE')),
        # Vignettes physiques (sur les véhicules ayant un statut actuel)
        physique_attribue=Count('id', filter=Q(statut_actuel__statut_physique='ATTRIBUE')),
        physique_non_attribue=Count('id', filter=Q(
            statut_actuel__statut_physique='NON_ATTRIBUE',
            statut_approbation='APPROUVE',
        )),
    )

    # Taux de recouvrement : véhicules VERT / total approuvés
    taux = 0
    if stats['approuves']:
        taux = round(stats['total_verts'] / stats['approuves'] * 100, 1)

    # Plaintes ouvertes
    plaintes_ouvertes = Plainte.objects.filter(statut__in=['OUVERTE', 'EN_COURS']).count()

    # Transferts en attente
    transferts_attente = DemandeTransfert.objects.filter(statut='EN_ATTENTE').count()

    return Response({
        **stats,
        'taux_recouvrement': taux,
        'plaintes_ouvertes': plaintes_ouvertes,
        'transferts_attente': transferts_attente,
    })


# =============== ADMIN — UTILISATEURS ===============

class AdminUtilisateurViewSet(viewsets.ModelViewSet):
    queryset = Utilisateur.objects.all().order_by('role', 'nom')
    permission_classes = [AdminOnlyPermission]

    def get_serializer_class(self):
        if self.action == 'create':
            return UtilisateurCreateSerializer
        if self.action in ['update', 'partial_update']:
            return UtilisateurUpdateSerializer
        return UtilisateurSerializer

    def perform_create(self, serializer):
        user = serializer.save()
        log_audit(self.request, CategorieAudit.UTILISATEUR,
                  f'Création compte utilisateur',
                  detail=f'Rôle: {user.role}',
                  objet_type='Utilisateur', objet_id=user.id, objet_label=user.email)

    def perform_update(self, serializer):
        user = serializer.save()
        log_audit(self.request, CategorieAudit.UTILISATEUR,
                  f'Modification compte utilisateur',
                  detail=f'Rôle: {user.role} | Actif: {user.est_actif}',
                  objet_type='Utilisateur', objet_id=user.id, objet_label=user.email)

    def perform_destroy(self, instance):
        log_audit(self.request, CategorieAudit.UTILISATEUR,
                  f'Suppression compte utilisateur',
                  detail=f'Rôle: {instance.role}',
                  objet_type='Utilisateur', objet_id=instance.id, objet_label=instance.email)
        instance.delete()

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        user = self.get_object()
        new_password = request.data.get('password', '').strip()
        if len(new_password) < 8:
            return Response({'error': 'Le mot de passe doit contenir au moins 8 caractères.'}, status=400)
        user.set_password(new_password)
        user.save(update_fields=['password'])
        log_audit(request, CategorieAudit.UTILISATEUR,
                  'Réinitialisation mot de passe',
                  objet_type='Utilisateur', objet_id=user.id, objet_label=user.email)
        return Response({'success': True, 'message': 'Mot de passe réinitialisé.'})

    @action(detail=False, methods=['post'])
    def change_my_password(self, request):
        ancien = request.data.get('ancien_password', '')
        nouveau = request.data.get('nouveau_password', '').strip()
        if not request.user.check_password(ancien):
            return Response({'error': 'Ancien mot de passe incorrect.'}, status=400)
        if len(nouveau) < 8:
            return Response({'error': 'Le nouveau mot de passe doit contenir au moins 8 caractères.'}, status=400)
        request.user.set_password(nouveau)
        request.user.save(update_fields=['password'])
        log_audit(request, CategorieAudit.UTILISATEUR,
                  'Changement de mot de passe personnel',
                  objet_type='Utilisateur', objet_id=request.user.id, objet_label=request.user.email)
        return Response({'success': True, 'message': 'Mot de passe modifié. Veuillez vous reconnecter.'})


# =============== ADMIN — PARAMÈTRES ===============

PARAMETRES_DEFAUT = [
    ('dashboard_roles',    'SUPERVISEUR_DGI,ADMINISTRATEUR_SYSTEME',                         'Rôles pouvant accéder au dashboard'),
    ('vehicules_roles',    'AGENT_DGI,SUPERVISEUR_DGI,ADMINISTRATEUR_SYSTEME,POLICE',        'Rôles pouvant voir la liste des véhicules'),
    ('saisie_roles',       'AGENT_DGI,SUPERVISEUR_DGI,ADMINISTRATEUR_SYSTEME',               'Rôles pouvant enregistrer un véhicule'),
    ('verification_roles', 'POLICE,AGENT_DGI,SUPERVISEUR_DGI,ADMINISTRATEUR_SYSTEME',        'Rôles pouvant accéder à la vérification'),
    ('nom_plateforme',     'MY CAR',                                                          'Nom affiché de la plateforme'),
    ('nom_organisation',   'DGI — République du Niger',                                       'Nom de l\'organisation'),
    ('devise',             'FCFA',                                                             'Devise utilisée pour les montants'),
    ('duree_vignette_jours', '365',                                                            'Durée de validité d\'une vignette (en jours)'),
    ('otp_expiration_minutes', '10',                                                           'Durée de validité d\'un OTP (en minutes)'),
]

class ParametrePlateformeViewSet(viewsets.ModelViewSet):
    queryset = ParametrePlateforme.objects.all().order_by('cle')
    serializer_class = ParametrePlateformeSerializer
    permission_classes = [AdminOnlyPermission]

    @action(detail=False, methods=['get'])
    def publics(self, request):
        """Paramètres lisibles par tous les utilisateurs authentifiés (nom plateforme, devise…)."""
        cles_publiques = ['nom_plateforme', 'nom_organisation', 'devise']
        params = ParametrePlateforme.objects.filter(cle__in=cles_publiques)
        return Response({p.cle: p.valeur for p in params})


# =============== PORTAIL CONTRIBUABLE (PUBLIC) ===============

def _masquer_telephone(tel):
    """Ex: +22790000001 → +227 90••••01"""
    t = tel.replace(' ', '')
    if len(t) >= 4:
        return t[:-6] + '••••' + t[-2:]
    return '••••••'


@api_view(['POST'])
@permission_classes([])
def public_demander_otp(request):
    """
    Étape 1 : le propriétaire saisit son immatriculation.
    Le système envoie un OTP sur le téléphone enregistré en base.
    Données personnelles NON exposées — seul le téléphone masqué est retourné.
    """
    immat  = request.data.get('immatriculation', '').strip().upper()
    region = request.data.get('region', '').strip()

    if not immat:
        return Response({'error': 'Immatriculation requise.'}, status=400)

    qs = Automobile.objects.select_related('region', 'statut_actuel').filter(immatriculation__iexact=immat)
    if region:
        qs = qs.filter(region__nom_region__iexact=region)
    auto = qs.first()

    if not auto:
        return Response({'error': 'Véhicule introuvable. Vérifiez l\'immatriculation et la région.'}, status=404)

    # Invalider les OTP précédents non utilisés
    OTPVerification.objects.filter(automobile=auto, est_utilise=False).update(est_utilise=True)

    code = f"{random.randint(0, 999999):06d}"
    otp  = OTPVerification.objects.create(
        automobile=auto,
        code=code,
        date_expiration=timezone.now() + timedelta(minutes=10),
    )

    # Envoi SMS (mock en dev, réel en prod via SMS_PROVIDER)
    sms_result = sms_otp(auto.telephone, code, auto.immatriculation)

    resp = {
        'otp_id':           otp.id,
        'telephone_masque': _masquer_telephone(auto.telephone),
        'expires_in':       600,
        'sms_envoye':       sms_result['ok'],
    }
    # En démo uniquement : retourner le code en clair
    import os
    if os.getenv('MOCK_SERVICES', 'True').lower() in ('true', '1', 'yes'):
        resp['demo_code']   = code
        resp['demo_notice'] = 'Mode démo : SMS simulé. En production, le code est envoyé par SMS uniquement.'
    return Response(resp, status=200)


@api_view(['POST'])
@permission_classes([])
def public_verifier_otp(request):
    """
    Étape 2 : vérification du code OTP.
    Si valide → retourne un session_token (15 min) + les données du véhicule.
    """
    otp_id = request.data.get('otp_id')
    code   = request.data.get('code', '').strip()

    if not all([otp_id, code]):
        return Response({'error': 'otp_id et code requis.'}, status=400)

    try:
        otp = OTPVerification.objects.select_related('automobile__region', 'automobile__statut_actuel').get(id=otp_id)
    except OTPVerification.DoesNotExist:
        return Response({'error': 'Session OTP introuvable.'}, status=404)

    if not otp.est_valide():
        return Response({'error': 'Code expiré ou déjà utilisé. Veuillez recommencer.'}, status=400)

    if otp.code != code:
        return Response({'error': 'Code incorrect. Vérifiez le SMS reçu.'}, status=400)

    otp.est_utilise = True
    otp.save(update_fields=['est_utilise'])

    auto = otp.automobile
    s    = auto.statut_actuel

    return Response({
        'session_token': otp.session_token,
        'expires_in':    900,
        'vehicule': {
            'id':              auto.id,
            'immatriculation': auto.immatriculation,
            'region':          auto.region.nom_region,
            'proprietaire':    f"{auto.nom} {auto.prenom}",
            'marque':          f"{auto.marque} {auto.modele or ''}".strip(),
            'type_vehicule':   auto.type_vehicule,
            'montant_taxe':    float(auto.montant_taxe),
            'statut':          s.statut if s else 'ROUGE',
            'date_fin_validite': s.date_fin_validite.isoformat() if s else None,
        },
    })


@api_view(['POST'])
@permission_classes([])
def public_initier_paiement(request):
    """Initie un paiement mobile money (simulation). Requiert un session_token valide."""
    session_token = request.data.get('session_token', '').strip()
    operateur     = request.data.get('operateur')
    telephone     = request.data.get('telephone', '').strip()

    duree_annees  = int(request.data.get('duree_annees', 1))
    if duree_annees not in [1, 2, 3]:
        return Response({'error': 'duree_annees doit être 1, 2 ou 3.'}, status=400)

    if not all([session_token, operateur, telephone]):
        return Response({'error': 'Champs requis : session_token, operateur, telephone.'}, status=400)

    if operateur not in OperateurPaiement.values:
        return Response({'error': f'Opérateur invalide.'}, status=400)

    try:
        otp = OTPVerification.objects.select_related('automobile').get(
            session_token=session_token,
            est_utilise=True,
            date_expiration__gte=timezone.now() - timedelta(minutes=15),
        )
    except OTPVerification.DoesNotExist:
        return Response({'error': 'Session expirée ou invalide. Veuillez vous réauthentifier.'}, status=403)

    # Un seul paiement autorisé par session OTP
    if hasattr(otp, 'paiement'):
        return Response({'error': 'Un paiement a déjà été initié pour cette session. Veuillez recommencer.'}, status=400)

    auto = otp.automobile

    # Véhicule doit être approuvé
    if auto.statut_approbation != StatutApprobationVehicule.APPROUVE:
        return Response({
            'error': f'Ce véhicule n\'est pas approuvé (statut : {auto.statut_approbation}). '
                     f'Contactez la DGI pour faire valider votre véhicule.'
        }, status=403)

    # Paiement autorisé uniquement si statut ORANGE
    statut_vignette = auto.statut_actuel.statut if auto.statut_actuel else StatutVignetteChoix.ROUGE
    if statut_vignette != StatutVignetteChoix.ORANGE:
        messages = {
            StatutVignetteChoix.VERT:  'Votre vignette est encore valide (VERTE). Le renouvellement ne sera possible que lorsqu\'elle sera ORANGE.',
            StatutVignetteChoix.ROUGE: 'Votre vignette est expirée (ROUGE). Le renouvellement en ligne n\'est pas possible. Présentez-vous à la DGI.',
        }
        return Response({
            'error': messages.get(statut_vignette, f'Statut {statut_vignette} ne permet pas le paiement en ligne.')
        }, status=403)

    Paiement.objects.filter(automobile=auto, statut=StatutPaiement.EN_ATTENTE).update(statut=StatutPaiement.ECHOUE)

    paiement = Paiement.objects.create(
        automobile=auto,
        otp=otp,
        montant=auto.montant_taxe * duree_annees,
        duree_annees=duree_annees,
        operateur=operateur,
        telephone=telephone,
    )

    return Response({
        'reference': paiement.reference,
        'montant':   float(paiement.montant),
        'operateur': paiement.operateur,
        'telephone': paiement.telephone,
        'statut':    paiement.statut,
        'message':   f'Demande envoyée. Confirmez sur votre téléphone.',
    }, status=201)


@api_view(['POST'])
@permission_classes([])
def public_confirmer_paiement(request, reference):
    """Simule la confirmation de paiement par l'opérateur mobile money."""
    try:
        paiement = Paiement.objects.select_related('automobile__region').get(reference=reference)
    except Paiement.DoesNotExist:
        return Response({'error': 'Référence de paiement introuvable.'}, status=404)

    if paiement.statut == StatutPaiement.CONFIRME:
        return Response({'error': 'Ce paiement est déjà confirmé.'}, status=400)

    if paiement.statut == StatutPaiement.ECHOUE:
        return Response({'error': 'Ce paiement a échoué. Veuillez recommencer.'}, status=400)

    # Simulation : création de la nouvelle vignette VERT
    auto = paiement.automobile
    date_fin = timezone.now().date() + timedelta(days=365 * paiement.duree_annees)
    physique_precedent = auto.statut_actuel.statut_physique if auto.statut_actuel else StatutPhysiqueVignette.NON_ATTRIBUE
    sv = StatutVignette.objects.create(
        automobile=auto,
        statut=StatutVignetteChoix.VERT,
        date_debut_validite=timezone.now().date(),
        date_fin_validite=date_fin,
        type_modification='AUTOMATIQUE',
        operateur=None,
        mobile_payment_ref=paiement.reference,
        statut_physique=physique_precedent,
    )
    auto.statut_actuel = sv
    auto.save(update_fields=['statut_actuel'])

    paiement.statut           = StatutPaiement.CONFIRME
    paiement.date_confirmation = timezone.now()
    paiement.statut_vignette  = sv
    paiement.save()

    # Générer QR
    qr_data = auto.generer_qr_data()
    qr = qrcode.QRCode(version=10, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=8, border=3)
    qr.add_data(json.dumps(qr_data))
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    # SMS de confirmation
    sms_confirmation_paiement(
        auto.telephone, auto.immatriculation,
        float(paiement.montant), paiement.reference
    )

    HistoriqueConsultation.objects.create(
        utilisateur=None,
        automobile=auto,
        action_performee=f'Paiement confirmé | Durée: {paiement.duree_annees} an(s) | Réf: {paiement.reference}',
        ip_address=request.META.get('REMOTE_ADDR'),
    )
    JournalAudit.objects.create(
        utilisateur=None,
        utilisateur_email='portail-contribuable',
        utilisateur_role='CONTRIBUABLE',
        categorie=CategorieAudit.PAIEMENT,
        action='Paiement vignette confirmé',
        detail=f'Réf: {paiement.reference} | Montant: {paiement.montant} FCFA | Durée: {paiement.duree_annees} an(s)',
        objet_type='Automobile', objet_id=auto.id, objet_label=auto.immatriculation,
        ip_address=request.META.get('REMOTE_ADDR'),
    )

    return Response({
        'statut':          'CONFIRME',
        'reference':       paiement.reference,
        'montant':         float(paiement.montant),
        'operateur':       paiement.operateur,
        'date_confirmation': paiement.date_confirmation.isoformat(),
        'vignette': {
            'statut':            sv.statut,
            'date_debut':        sv.date_debut_validite.isoformat(),
            'date_fin':          sv.date_fin_validite.isoformat(),
            'code_securite':     sv.code_securite,
        },
        'qr_code_base64': qr_b64,
        'qr_data':        qr_data,
    })


# =============== DASHBOARD FINANCIER ===============

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_financier(request):
    """
    Agrégation des paiements confirmés avec filtres :
    date_debut, date_fin, region, agent (utilisateur_id)
    """
    from django.db.models.functions import TruncMonth, TruncDay
    from django.db.models import Sum, Count, Q
    import csv
    from django.http import HttpResponse as DjangoHttpResponse

    qs = Paiement.objects.filter(statut=StatutPaiement.CONFIRME).select_related(
        'automobile__region', 'otp'
    )

    # Agent DGI : restreint à sa région
    if request.user.role == RoleUtilisateur.AGENT_DGI and request.user.region_id:
        qs = qs.filter(automobile__region_id=request.user.region_id)

    # Filtres
    date_debut = request.query_params.get('date_debut')
    date_fin   = request.query_params.get('date_fin')
    region     = request.query_params.get('region')
    operateur  = request.query_params.get('operateur')

    if date_debut:
        qs = qs.filter(date_confirmation__date__gte=date_debut)
    if date_fin:
        qs = qs.filter(date_confirmation__date__lte=date_fin)
    if region:
        qs = qs.filter(automobile__region__nom_region__iexact=region)
    if operateur:
        qs = qs.filter(operateur=operateur)

    # Totaux globaux
    totaux = qs.aggregate(
        total_recettes=Sum('montant'),
        total_paiements=Count('id'),
    )

    # Par mois
    par_mois = list(
        qs.annotate(mois=TruncMonth('date_confirmation'))
          .values('mois')
          .annotate(recettes=Sum('montant'), nb=Count('id'))
          .order_by('mois')
    )

    # Par région
    par_region = list(
        qs.values('automobile__region__nom_region')
          .annotate(recettes=Sum('montant'), nb=Count('id'))
          .order_by('-recettes')
    )

    # Par opérateur mobile money
    par_operateur = list(
        qs.values('operateur')
          .annotate(recettes=Sum('montant'), nb=Count('id'))
          .order_by('-recettes')
    )

    # Derniers paiements
    derniers = list(
        qs.order_by('-date_confirmation')[:20]
          .values(
              'reference', 'montant', 'operateur', 'date_confirmation',
              'automobile__immatriculation', 'automobile__region__nom_region',
          )
    )

    # Export CSV
    if request.query_params.get('export') == 'csv':
        response = DjangoHttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="recettes_mycar.csv"'
        writer = csv.writer(response)
        writer.writerow(['Référence', 'Immatriculation', 'Région', 'Opérateur', 'Montant (FCFA)', 'Date'])
        for p in qs.order_by('-date_confirmation').values(
            'reference', 'automobile__immatriculation',
            'automobile__region__nom_region', 'operateur', 'montant', 'date_confirmation'
        ):
            writer.writerow([
                p['reference'],
                p['automobile__immatriculation'],
                p['automobile__region__nom_region'],
                p['operateur'],
                p['montant'],
                p['date_confirmation'].strftime('%d/%m/%Y %H:%M') if p['date_confirmation'] else '',
            ])
        return response

    return Response({
        'totaux': {
            'recettes':   float(totaux['total_recettes'] or 0),
            'paiements':  totaux['total_paiements'],
        },
        'par_mois': [
            {'mois': e['mois'].strftime('%Y-%m'), 'recettes': float(e['recettes']), 'nb': e['nb']}
            for e in par_mois
        ],
        'par_region': [
            {'region': e['automobile__region__nom_region'], 'recettes': float(e['recettes']), 'nb': e['nb']}
            for e in par_region
        ],
        'par_operateur': [
            {'operateur': e['operateur'], 'recettes': float(e['recettes']), 'nb': e['nb']}
            for e in par_operateur
        ],
        'derniers_paiements': [
            {
                'reference':      p['reference'],
                'immatriculation': p['automobile__immatriculation'],
                'region':         p['automobile__region__nom_region'],
                'operateur':      p['operateur'],
                'montant':        float(p['montant']),
                'date':           p['date_confirmation'].strftime('%d/%m/%Y %H:%M') if p['date_confirmation'] else '',
            }
            for p in derniers
        ],
    })


# =============== AGENT DE DISTRIBUTION — ATTRIBUTION VIGNETTE PHYSIQUE ===============

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def agent_rechercher_vehicule(request):
    """
    Recherche d'un véhicule par immatriculation (Agent de distribution).
    Retourne les infos nécessaires à la remise physique de vignette.
    """
    roles_autorises = {RoleUtilisateur.AGENT_DISTRIB, RoleUtilisateur.AGENT_DGI, RoleUtilisateur.ADMIN_SYS, RoleUtilisateur.SUP_DGI}
    if request.user.role not in roles_autorises:
        return Response({'error': 'Accès non autorisé.'}, status=403)

    immat = request.query_params.get('immatriculation', '').strip().upper()
    if not immat:
        return Response({'error': 'Paramètre immatriculation requis.'}, status=400)

    auto = Automobile.objects.select_related('region', 'statut_actuel').filter(
        immatriculation__iexact=immat
    ).first()

    if not auto:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    s = auto.statut_actuel
    return Response({
        'id': auto.id,
        'immatriculation': auto.immatriculation,
        'region': auto.region.nom_region,
        'proprietaire': f"{auto.nom} {auto.prenom}",
        'telephone': auto.telephone,
        'marque': f"{auto.marque} {auto.modele or ''}".strip(),
        'montant_taxe': float(auto.montant_taxe),
        'statut_approbation': auto.statut_approbation,
        'statut_vignette': s.statut if s else 'ROUGE',
        'statut_physique': s.statut_physique if s else 'NON_ATTRIBUE',
        'date_fin_validite': s.date_fin_validite.isoformat() if s else None,
        'paiement_confirme': Paiement.objects.filter(
            automobile=auto, statut=StatutPaiement.CONFIRME
        ).order_by('-date_confirmation').first() is not None,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def agent_paiement_agence(request, automobile_id):
    """
    Paiement en agence : l'agent règle la vignette pour le compte du contribuable.
    Pas d'OTP requis — l'agent est physiquement présent et authentifié.
    """
    roles_autorises = {RoleUtilisateur.AGENT_DISTRIB, RoleUtilisateur.AGENT_DGI, RoleUtilisateur.ADMIN_SYS}
    if request.user.role not in roles_autorises:
        return Response({'error': 'Accès non autorisé.'}, status=403)

    try:
        auto = Automobile.objects.select_related('statut_actuel', 'region').get(id=automobile_id)
    except Automobile.DoesNotExist:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    if auto.statut_approbation != StatutApprobationVehicule.APPROUVE:
        return Response({'error': 'Ce véhicule n\'est pas approuvé.'}, status=400)

    if Paiement.objects.filter(automobile=auto, statut=StatutPaiement.CONFIRME).exists():
        return Response({'error': 'Un paiement est déjà confirmé pour ce véhicule.'}, status=400)

    operateur    = request.data.get('operateur', '').strip().upper()
    telephone    = request.data.get('telephone', '').strip()
    duree_annees = int(request.data.get('duree_annees', 1))

    if operateur not in OperateurPaiement.values:
        return Response({'error': f'Opérateur invalide. Valeurs : {OperateurPaiement.values}'}, status=400)
    if not telephone:
        return Response({'error': 'Le numéro de téléphone est obligatoire.'}, status=400)
    if duree_annees not in [1, 2, 3]:
        return Response({'error': 'duree_annees doit être 1, 2 ou 3.'}, status=400)

    montant = auto.montant_taxe * duree_annees
    date_fin = timezone.now().date() + timedelta(days=365 * duree_annees)

    # Annuler les paiements EN_ATTENTE existants
    Paiement.objects.filter(automobile=auto, statut=StatutPaiement.EN_ATTENTE).update(statut=StatutPaiement.ECHOUE)

    # Créer et confirmer immédiatement le paiement
    paiement = Paiement.objects.create(
        automobile=auto,
        montant=montant,
        duree_annees=duree_annees,
        operateur=operateur,
        telephone=telephone,
        statut=StatutPaiement.CONFIRME,
        date_confirmation=timezone.now(),
    )

    # Créer la vignette VERT (préserver l'attribution physique si déjà faite)
    physique_precedent = auto.statut_actuel.statut_physique if auto.statut_actuel else StatutPhysiqueVignette.NON_ATTRIBUE
    sv = StatutVignette.objects.create(
        automobile=auto,
        statut=StatutVignetteChoix.VERT,
        date_debut_validite=timezone.now().date(),
        date_fin_validite=date_fin,
        type_modification=TypeModification.MANUELLE,
        operateur=request.user,
        notes_admin=f'Paiement en agence par {request.user.nom} {request.user.prenom}',
        mobile_payment_ref=paiement.reference,
        statut_physique=physique_precedent,
    )
    auto.statut_actuel = sv
    auto.save(update_fields=['statut_actuel'])
    paiement.statut_vignette = sv
    paiement.save(update_fields=['statut_vignette'])

    sms_confirmation_paiement(auto.telephone, auto.immatriculation, float(montant), paiement.reference)

    log_audit(request, CategorieAudit.PAIEMENT,
              'Paiement en agence pour le compte du contribuable',
              detail=f'Montant: {montant} FCFA | Durée: {duree_annees} an(s) | Opérateur: {operateur} | Tél: {telephone}',
              objet_type='Automobile', objet_id=auto.id, objet_label=auto.immatriculation)

    return Response({
        'success':    True,
        'reference':  paiement.reference,
        'montant':    float(montant),
        'duree_annees': duree_annees,
        'date_fin':   date_fin.isoformat(),
        'statut_vignette': sv.statut,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def agent_attribuer_vignette(request, automobile_id):
    """
    Attribution de la vignette physique à un véhicule (Agent de distribution).
    Conditions : véhicule APPROUVE + paiement CONFIRME + statut physique NON_ATTRIBUE.
    """
    roles_autorises = {RoleUtilisateur.AGENT_DISTRIB, RoleUtilisateur.ADMIN_SYS}
    if request.user.role not in roles_autorises:
        return Response({'error': 'Accès non autorisé.'}, status=403)

    try:
        auto = Automobile.objects.select_related('statut_actuel').get(id=automobile_id)
    except Automobile.DoesNotExist:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    if auto.statut_approbation != StatutApprobationVehicule.APPROUVE:
        return Response({'error': 'Ce véhicule n\'est pas approuvé par l\'administrateur.'}, status=403)

    if not auto.statut_actuel:
        return Response({'error': 'Aucune vignette numérique active pour ce véhicule.'}, status=400)

    if auto.statut_actuel.statut_physique == StatutPhysiqueVignette.ATTRIBUE:
        return Response({'error': 'Une vignette physique a déjà été attribuée à ce véhicule.'}, status=400)

    paiement_ok = Paiement.objects.filter(
        automobile=auto, statut=StatutPaiement.CONFIRME
    ).exists()
    if not paiement_ok:
        return Response({'error': 'Aucun paiement confirmé pour ce véhicule.'}, status=403)

    sv = auto.statut_actuel
    sv.statut_physique = StatutPhysiqueVignette.ATTRIBUE
    sv.notes_admin = (sv.notes_admin + '\n' if sv.notes_admin else '') + \
        f'Vignette physique attribuée par {request.user.nom} {request.user.prenom} le {timezone.now().strftime("%d/%m/%Y %H:%M")}'
    sv.save(update_fields=['statut_physique', 'notes_admin'])

    HistoriqueConsultation.objects.create(
        utilisateur=request.user,
        automobile=auto,
        action_performee=f'Attribution vignette physique par Agent {request.user.nom} {request.user.prenom}',
        ip_address=request.META.get('REMOTE_ADDR'),
    )
    log_audit(request, CategorieAudit.DISTRIBUTION,
              'Attribution vignette physique',
              objet_type='Automobile', objet_id=auto.id, objet_label=auto.immatriculation)

    return Response({
        'success': True,
        'immatriculation': auto.immatriculation,
        'statut_physique': StatutPhysiqueVignette.ATTRIBUE,
        'attribue_par': f"{request.user.nom} {request.user.prenom}",
        'date_attribution': timezone.now().isoformat(),
    })


# =============== TRANSFERT DE PROPRIÉTÉ ===============

class DemandeTransfertViewSet(viewsets.ModelViewSet):
    queryset = DemandeTransfert.objects.select_related('automobile', 'traite_par').order_by('-date_demande')
    serializer_class = DemandeTransfertSerializer

    def get_permissions(self):
        if self.action == 'create':
            return [IsAuthenticated()]
        return [IsAuthenticated()]

    def perform_create(self, request_data, auto):
        return DemandeTransfert.objects.create(
            automobile=auto,
            ancien_nom=auto.nom,
            ancien_prenom=auto.prenom,
            ancien_telephone=auto.telephone,
            **request_data,
        )

    def create(self, request, *args, **kwargs):
        automobile_id = request.data.get('automobile')
        try:
            auto = Automobile.objects.get(id=automobile_id)
        except Automobile.DoesNotExist:
            return Response({'error': 'Véhicule introuvable.'}, status=404)

        data = {
            'nouveau_nom':       request.data.get('nouveau_nom', '').strip(),
            'nouveau_prenom':    request.data.get('nouveau_prenom', '').strip(),
            'nouveau_telephone': request.data.get('nouveau_telephone', '').strip(),
            'motif':             request.data.get('motif', '').strip(),
        }
        if not all([data['nouveau_nom'], data['nouveau_prenom'], data['nouveau_telephone']]):
            return Response({'error': 'Champs obligatoires : nouveau_nom, nouveau_prenom, nouveau_telephone'}, status=400)

        demande = self.perform_create(data, auto)
        serializer = self.get_serializer(demande)
        return Response(serializer.data, status=201)

    @action(detail=True, methods=['post'])
    def traiter(self, request, pk=None):
        """Approuver ou rejeter un transfert (Admin système uniquement)."""
        if request.user.role != RoleUtilisateur.ADMIN_SYS:
            return Response({'error': 'Seul l\'administrateur système peut traiter les transferts.'}, status=403)

        demande = self.get_object()
        if demande.statut != StatutTransfert.EN_ATTENTE:
            return Response({'error': f'Cette demande est déjà traitée ({demande.statut}).'}, status=400)

        decision   = request.data.get('decision', '').strip().upper()
        notes_admin = request.data.get('notes', '').strip()

        if decision not in [StatutTransfert.APPROUVE, StatutTransfert.REJETE]:
            return Response({'error': 'decision doit être APPROUVE ou REJETE.'}, status=400)

        demande.statut          = decision
        demande.notes_admin     = notes_admin
        demande.traite_par      = request.user
        demande.date_traitement = timezone.now()
        demande.save()

        if decision == StatutTransfert.APPROUVE:
            # Mettre à jour le propriétaire du véhicule
            auto = demande.automobile
            auto.nom       = demande.nouveau_nom
            auto.prenom    = demande.nouveau_prenom
            auto.telephone = demande.nouveau_telephone
            auto.save(update_fields=['nom', 'prenom', 'telephone'])

            HistoriqueConsultation.objects.create(
                utilisateur=request.user,
                automobile=auto,
                action_performee=f'Transfert propriété approuvé : {demande.ancien_nom} → {demande.nouveau_nom} {demande.nouveau_prenom}',
                ip_address=request.META.get('REMOTE_ADDR'),
            )
            # Notifier le nouveau propriétaire
            sms_transfert_approuve(demande.nouveau_telephone, auto.immatriculation)

        log_audit(request, CategorieAudit.TRANSFERT,
                  f'Transfert de propriété : {decision}',
                  detail=f'{demande.ancien_nom} → {demande.nouveau_nom} {demande.nouveau_prenom}',
                  objet_type='DemandeTransfert', objet_id=demande.id, objet_label=demande.automobile.immatriculation)

        return Response({
            'success': True,
            'decision': decision,
            'automobile': demande.automobile.immatriculation,
        })


# =============== CERTIFICAT / REÇU PDF (HTML print-ready) ===============

from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def certificat_vignette(request, automobile_id):
    """Génère les données pour le certificat numérique (QR + infos vignette). Réservé à l'admin système."""
    if request.user.role != RoleUtilisateur.ADMIN_SYS:
        return Response({'error': 'Le certificat numérique est réservé à l\'administrateur système (CARBOOK).'}, status=403)

    auto = get_object_or_404(Automobile, id=automobile_id)

    if not auto.statut_actuel:
        return Response({'error': 'Aucune vignette active pour ce véhicule.'}, status=404)

    # QR code
    qr_data = auto.generer_qr_data()
    qr = qrcode.QRCode(version=10, error_correction=qrcode.constants.ERROR_CORRECT_H, box_size=8, border=3)
    qr.add_data(json.dumps(qr_data))
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    qr_b64 = base64.b64encode(buf.getvalue()).decode()

    # Dernier paiement confirmé
    dernier_paiement = Paiement.objects.filter(
        automobile=auto, statut=StatutPaiement.CONFIRME
    ).order_by('-date_confirmation').first()

    sv = auto.statut_actuel
    return Response({
        'vehicule': {
            'immatriculation': auto.immatriculation,
            'region': auto.region.nom_region,
            'proprietaire': f"{auto.nom} {auto.prenom}",
            'telephone': auto.telephone,
            'marque': f"{auto.marque} {auto.modele or ''}".strip(),
            'type_vehicule': auto.type_vehicule,
            'numero_chassis': auto.numero_chassis,
            'energie': auto.energie,
            'puissance_cv': auto.puissance_cv,
            'annee_fabrication': auto.annee_fabrication,
        },
        'vignette': {
            'statut': sv.statut,
            'statut_physique': sv.statut_physique,
            'date_debut': sv.date_debut_validite.isoformat(),
            'date_fin': sv.date_fin_validite.isoformat(),
            'code_securite': sv.code_securite,
            'type_modification': sv.type_modification,
        },
        'paiement': {
            'reference': dernier_paiement.reference if dernier_paiement else None,
            'montant': float(dernier_paiement.montant) if dernier_paiement else float(auto.montant_taxe),
            'operateur': dernier_paiement.operateur if dernier_paiement else None,
            'date': dernier_paiement.date_confirmation.isoformat() if dernier_paiement and dernier_paiement.date_confirmation else None,
        },
        'qr_code_base64': qr_b64,
        'genere_le': timezone.now().isoformat(),
    })


# =============== PLAINTES / LITIGES ===============

class PlainteViewSet(viewsets.ModelViewSet):
    queryset = Plainte.objects.select_related('automobile', 'traite_par').order_by('-date_creation')
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        from .serializers import PlainteSerializer
        return PlainteSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        statut = self.request.query_params.get('statut')
        if statut:
            qs = qs.filter(statut=statut)
        return qs

    def perform_create(self, serializer):
        plainte = serializer.save()
        # SMS de confirmation au plaignant
        sms_plainte_recue(plainte.telephone, plainte.reference)

    @action(detail=True, methods=['post'])
    def traiter(self, request, pk=None):
        """Répondre / clore une plainte (Admin ou Superviseur DGI)."""
        roles_autorises = {RoleUtilisateur.ADMIN_SYS, RoleUtilisateur.SUP_DGI}
        if request.user.role not in roles_autorises:
            return Response({'error': 'Accès non autorisé.'}, status=403)

        plainte = self.get_object()
        nouveau_statut = request.data.get('statut', StatutPlainte.RESOLUE)
        reponse        = request.data.get('reponse', '').strip()

        if nouveau_statut not in StatutPlainte.values:
            return Response({'error': f'Statut invalide. Valeurs : {StatutPlainte.values}'}, status=400)

        plainte.statut          = nouveau_statut
        plainte.reponse_admin   = reponse
        plainte.traite_par      = request.user
        plainte.date_traitement = timezone.now()
        plainte.save()

        return Response({'success': True, 'reference': plainte.reference, 'statut': plainte.statut})


# =============== EXPORT EXCEL ===============

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_excel(request):
    """
    Export Excel (.xlsx) des véhicules ou des paiements.
    ?type=vehicules|paiements (défaut: vehicules)
    Filtres : region, statut, date_debut, date_fin
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    type_export = request.query_params.get('type', 'vehicules')

    wb = Workbook()
    ws = wb.active

    # Styles
    header_font  = Font(bold=True, color='FFFFFF', size=11)
    header_fill  = PatternFill(fill_type='solid', fgColor='0F766E')
    center       = Alignment(horizontal='center', vertical='center')
    thin_border  = Border(
        bottom=Side(style='thin', color='E5E7EB'),
    )

    def style_header(row_num):
        for cell in ws[row_num]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center

    def style_row(row_num, alternate=False):
        fill_color = 'F0FDF4' if alternate else 'FFFFFF'
        for cell in ws[row_num]:
            cell.fill      = PatternFill(fill_type='solid', fgColor=fill_color)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical='center')

    if type_export == 'vehicules':
        ws.title = 'Véhicules'
        qs = Automobile.objects.select_related('region', 'statut_actuel').all()

        region = request.query_params.get('region')
        statut = request.query_params.get('statut')
        if region: qs = qs.filter(region__nom_region__iexact=region)
        if statut: qs = qs.filter(statut_actuel__statut__iexact=statut)

        headers = [
            'Immatriculation', 'Région', 'Propriétaire', 'Téléphone',
            'Type', 'Marque', 'Énergie', 'Puissance (CV)', 'Châssis',
            'Approbation', 'Statut vignette', 'Fin validité', 'Statut physique',
            'Date création',
        ]
        ws.append(headers)
        style_header(1)
        ws.row_dimensions[1].height = 20

        for i, auto in enumerate(qs, start=2):
            s = auto.statut_actuel
            ws.append([
                auto.immatriculation,
                auto.region.nom_region,
                f"{auto.nom} {auto.prenom}",
                auto.telephone,
                auto.type_vehicule,
                f"{auto.marque} {auto.modele or ''}".strip(),
                auto.energie or '',
                auto.puissance_cv,
                auto.numero_chassis,
                auto.statut_approbation,
                s.statut if s else 'ROUGE',
                s.date_fin_validite.strftime('%d/%m/%Y') if s else '',
                s.statut_physique if s else '',
                auto.date_creation.strftime('%d/%m/%Y'),
            ])
            style_row(i, alternate=(i % 2 == 0))

        col_widths = [15, 12, 22, 16, 10, 18, 10, 12, 20, 12, 14, 12, 14, 14]

    elif type_export == 'paiements':
        ws.title = 'Paiements'
        qs = Paiement.objects.filter(statut=StatutPaiement.CONFIRME).select_related(
            'automobile__region'
        ).order_by('-date_confirmation')

        region    = request.query_params.get('region')
        operateur = request.query_params.get('operateur')
        date_debut = request.query_params.get('date_debut')
        date_fin   = request.query_params.get('date_fin')

        if region:     qs = qs.filter(automobile__region__nom_region__iexact=region)
        if operateur:  qs = qs.filter(operateur=operateur)
        if date_debut: qs = qs.filter(date_confirmation__date__gte=date_debut)
        if date_fin:   qs = qs.filter(date_confirmation__date__lte=date_fin)

        headers = [
            'Référence', 'Immatriculation', 'Région', 'Propriétaire',
            'Opérateur', 'Téléphone', 'Montant (FCFA)', 'Date confirmation',
        ]
        ws.append(headers)
        style_header(1)
        ws.row_dimensions[1].height = 20

        for i, p in enumerate(qs, start=2):
            ws.append([
                p.reference,
                p.automobile.immatriculation,
                p.automobile.region.nom_region,
                f"{p.automobile.nom} {p.automobile.prenom}",
                p.operateur,
                p.telephone,
                float(p.montant),
                p.date_confirmation.strftime('%d/%m/%Y %H:%M') if p.date_confirmation else '',
            ])
            style_row(i, alternate=(i % 2 == 0))

        col_widths = [18, 15, 12, 22, 14, 16, 16, 18]

    else:
        return Response({'error': 'type doit être vehicules ou paiements'}, status=400)

    # Largeurs colonnes
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # Figer la ligne d'en-tête
    ws.freeze_panes = 'A2'

    # Retourner le fichier
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"mycar_{type_export}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# =============== HEALTH CHECK ===============

@api_view(['GET'])
def health_check(request):
    return Response({'status': 'ok', 'timestamp': timezone.now()}, status=200)


# =============== INSPECTION GPS (POLICE) ===============

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def log_inspection_gps(request):
    """
    Enregistre une inspection police avec coordonnées GPS.
    Body: { immatriculation, region, latitude, longitude }
    """
    roles_autorises = {RoleUtilisateur.POLICE, RoleUtilisateur.ADMIN_SYS, RoleUtilisateur.SUP_DGI, RoleUtilisateur.AGENT_DGI}
    if request.user.role not in roles_autorises:
        return Response({'error': 'Accès non autorisé.'}, status=403)

    immat    = request.data.get('immatriculation', '').strip().upper()
    region   = request.data.get('region', '').strip()
    latitude = request.data.get('latitude')
    longitude = request.data.get('longitude')

    auto = Automobile.objects.filter(immatriculation__iexact=immat).first()
    if not auto:
        return Response({'error': 'Véhicule introuvable.'}, status=404)

    gps_info = ''
    if latitude is not None and longitude is not None:
        gps_info = f' | GPS: {latitude},{longitude}'

    HistoriqueConsultation.objects.create(
        utilisateur=request.user,
        automobile=auto,
        action_performee=f'Inspection police{gps_info}',
        ip_address=request.META.get('REMOTE_ADDR'),
    )
    return Response({'ok': True})
